from openpyxl import load_workbook
from openpyxl import Workbook
import os

# the number to be searched
number = input('Enter number to find: ')
# column to be searched
x = int(input('Enter containing column number: '))
x -= 1

# directory containing the workbooks
directory = "D:\parse"

wb1 = Workbook(write_only=True)
ws1 = wb1.create_sheet()

for root,dirs,files in os.walk(directory):
    for file in files:
        if file.endswith(".xlsx"):
            wb2 = load_workbook(filename=file, read_only=True)
            ws2 = wb2.active

            for row in ws2.rows:
                list = []
                # change row[1] to row[x] where x is the column where the data to be matched is present
                if str(row[1].value) == number:
                    for cell in row:
                        list.append(cell.value)
                    ws1.append(list)

wb1.save('output.xlsx')
